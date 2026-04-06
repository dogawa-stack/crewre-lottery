#!/usr/bin/env python3
# lottery.py - crewre popupイベント抽選システム

import csv
import os
import random
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ==============================
# 設定（毎回変更する部分）
# ==============================

PAPERFORM_CSV = '/Users/ogawadaiki/Downloads/crewre 26 spring popup event (1).csv'
SHOPIFY_CSV   = '/Users/ogawadaiki/Downloads/customers_export/customers_export.csv'
OUTPUT_DIR    = '/Users/ogawadaiki/Downloads/'

# Google DriveファイルID（固定・上書き更新）
GDRIVE_FILE_ID = '1mhCKVPSDlCuZV4ER_O8tcckM7xUohMekFhLz9nrceIQ'

# 各枠の定員
CAPACITY_PER_SLOT = 10

# 各枠のVIP/潜在/新規 配分（合計=CAPACITY_PER_SLOT）
ALLOCATION = {'VIP': 3, '潜在': 3, '新規': 4}

# ステータス分類の閾値（毎回調整可能）
VIP_THRESHOLD        = 200000   # VIP: 合算購入額200,000円以上
POTENTIAL_REG_BEFORE = '2025'   # 潜在: EC-CUBE登録が2025年より前
# 新規: 2025年以降登録 または EC-CUBEに登録なし

# ==============================
# 枠定義（5/9・5/10 各16枠）
# ==============================

SLOT_NAMES  = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩','⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲','⑳']

# デフォルト設定（app.pyから上書き可能）
DEFAULT_DAYS = ['5/9(土)', '5/10(日)']
DEFAULT_SLOT_TIMES = [
    '10:00〜11:00','10:30〜11:30','11:00〜12:00','11:30〜12:30',
    '12:00〜13:00','12:30〜13:30','13:00〜14:00','13:30〜14:30',
    '14:00〜15:00','14:30〜15:30','15:00〜16:00','15:30〜16:30',
    '16:00〜17:00','16:30〜17:30','17:00〜18:00','18:00〜19:00',
]


def _normalize_slot_times(days, slot_times_per_day):
    """slot_times_per_day を日ごとのリストに正規化。
    slot_times_per_day が単一リストの場合は全日共通として扱う。"""
    if slot_times_per_day and isinstance(slot_times_per_day[0], str):
        # 全日共通
        return [slot_times_per_day] * len(days)
    return slot_times_per_day


def build_slot_defs(days, slot_times_per_day):
    """日程×時間帯からSLOT_DEFSを生成（日ごとに異なる時間帯に対応）"""
    per_day = _normalize_slot_times(days, slot_times_per_day)
    defs = {}
    sid = 1
    for d, day in enumerate(days):
        for i, t in enumerate(per_day[d]):
            name = SLOT_NAMES[i] if i < len(SLOT_NAMES) else f'({i+1})'
            defs[sid] = f'{day} {name} {t}'
            sid += 1
    return defs


def build_time_zone_map(days, slot_times_per_day):
    """日×時間グループ（午前/午後/夕方）でPaperformの選択肢マップを自動生成"""
    per_day = _normalize_slot_times(days, slot_times_per_day)

    def parse_start(t):
        start = t.split('〜')[0].strip()
        h, m = map(int, start.split(':'))
        return h * 60 + m

    def get_group(t):
        mins = parse_start(t)
        if mins < 12 * 60:
            return '午前'
        elif mins < 16 * 60:
            return '午後'
        else:
            return '夕方'

    zone_labels = ['①','②','③','④','⑤','⑥','⑦','⑧','⑨','⑩','⑪','⑫','⑬','⑭','⑮','⑯','⑰','⑱','⑲','⑳']
    zone_map = {}
    zone_idx = 0
    sid = 1
    all_slots = []

    for d, day in enumerate(days):
        groups = {}
        for t in per_day[d]:
            g = get_group(t)
            groups.setdefault(g, []).append(sid)
            all_slots.append(sid)
            sid += 1
        for g in ['午前', '午後', '夕方']:
            if g in groups:
                zone_map[zone_labels[zone_idx]] = groups[g]
                zone_idx += 1

    # いずれでも可
    zone_map[zone_labels[zone_idx]] = all_slots
    return zone_map


# モジュールレベルの変数（app.pyから一時的に上書きされる）
SLOT_DEFS     = build_slot_defs(DEFAULT_DAYS, DEFAULT_SLOT_TIMES)
TIME_ZONE_MAP = build_time_zone_map(DEFAULT_DAYS, DEFAULT_SLOT_TIMES)


# ==============================
# メイン処理
# ==============================

def classify(total_spent, eccube_registered=''):
    """購入金額・EC-CUBE登録時期からステータスを返す"""
    try:
        amount = float(str(total_spent).replace(',', '').replace("'", ''))
    except:
        amount = 0
    if amount >= VIP_THRESHOLD:
        return 'VIP'
    # EC-CUBEに登録あり かつ POTENTIAL_REG_BEFORE より前 → 潜在
    if eccube_registered and eccube_registered[:4] < POTENTIAL_REG_BEFORE:
        return '潜在'
    return '新規'


ECCUBE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'eccube_members.xlsx')

def load_eccube():
    """EC-CUBE XLSX → {email: {spent, registered}} の辞書"""
    eccube = {}
    if not os.path.exists(ECCUBE_PATH):
        return eccube
    wb = openpyxl.load_workbook(ECCUBE_PATH, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    email_idx = headers.index('E-MAIL') + 1
    spent_idx = headers.index('お買い上げ合計額') + 1
    date_idx  = headers.index('登録日') + 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[email_idx - 1]
        spent = row[spent_idx - 1]
        reg   = row[date_idx - 1]
        if email:
            try:
                s = float(str(spent).replace(',', '') or 0)
            except:
                s = 0
            reg_str = str(reg)[:7] if reg else ''  # 'YYYY-MM'
            eccube[str(email).strip().lower()] = {'spent': s, 'registered': reg_str}
    wb.close()
    return eccube


def load_shopify(path):
    """Shopify CSV → {email: {name, spent, eccube_registered}} の辞書（EC-CUBEデータと合算）"""
    eccube = load_eccube()
    customers = {}
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            clean = {k.strip(): v for k, v in row.items() if k}
            email = _find_col(clean, ['Email', 'メールアドレス']).strip().lower()
            if email:
                try:
                    shopify_spent = float(str(_find_col(clean, ['Total Spent', 'Amount spent'])).replace(',', '') or 0)
                except:
                    shopify_spent = 0
                ec = eccube.get(email, {})
                total = shopify_spent + ec.get('spent', 0)
                customers[email] = {
                    'name': f"{_find_col(clean, ['First Name'])} {_find_col(clean, ['Last Name'])}".strip(),
                    'spent': total,
                    'eccube_registered': ec.get('registered', ''),
                }
    # EC-CUBEにいてShopifyにいない会員も追加
    for email, ec in eccube.items():
        if email not in customers:
            customers[email] = {
                'name': '',
                'spent': ec['spent'],
                'eccube_registered': ec.get('registered', ''),
            }
    return customers


def parse_preferred_slots(slot_str):
    """希望日時文字列 → 対応する枠IDのセットを返す"""
    slot_ids = set()
    for key, ids in TIME_ZONE_MAP.items():
        if key in slot_str:
            slot_ids.update(ids)
    return slot_ids


def _find_col(row, candidates):
    """複数の候補列名から最初に見つかった列の値を返す。"""
    keys = list(row.keys())
    # 完全一致
    for c in candidates:
        if c in row:
            return row[c]
    # 前方一致を優先（キーが候補で始まるか）
    for c in candidates:
        for k in keys:
            if k and k.startswith(c):
                return row[k]
    # 部分一致（フォールバック）
    for c in candidates:
        for k in keys:
            if k and c in k:
                return row[k]
    return ''

def load_applicants(path, shopify):
    """Paperform CSV → 応募者リストを返す"""
    applicants = []
    # BOM対応: utf-8-sig はBOMがあれば除去、なくてもOK
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 列名の空白・不可視文字を除去して検索
            clean = {k.strip(): v for k, v in row.items() if k}

            email = _find_col(clean, ['メールアドレス', 'メール', 'Email']).strip().lower()
            # 姓名が別列の場合（姓=「お名前(...)」、名=空白列名）を結合
            last_name = _find_col(clean, ['氏名', 'お名前']).strip()
            # 名前列: 元キーが' '(スペース)→strip後''、どちらでも取得
            first_name = (clean.get('', '') or row.get(' ', '') or '').strip()
            name = f"{last_name} {first_name}".strip() if first_name else last_name
            slot_str = _find_col(clean, ['希望日時', 'ご希望のご来場日', '来場日']).strip()

            if not email or not name:
                continue

            preferred = parse_preferred_slots(slot_str)
            if not preferred:
                preferred = set(range(1, 33))  # 指定なし = 全枠対象

            customer = shopify.get(email, {})
            status = classify(customer.get('spent', 0), customer.get('eccube_registered', ''))

            # ペア参加判定（旧形式・新形式両対応）
            pair_col = _find_col(clean, ['ペア参加を希望しますか？', '参加形式をお選びください', '参加形式']).strip()
            is_pair = pair_col in ('希望する', 'ペア参加を希望する')

            # 同伴者（新形式: 試着不可の付き添い。定員カウント外）
            companion_name = _find_col(clean, ['同伴される方', '同伴者', '同伴者氏名']).strip()
            is_companion = pair_col == '同伴者様との参加を希望する' or bool(companion_name)

            applicants.append({
                'name':      name,
                'email':     email,
                'preferred': preferred,
                'status':    status,
                'slot_str':  slot_str,
                'is_pair':   is_pair,
                'pair_name': _find_col(clean, ['ペアの方', 'ペアの方の氏名', 'ペア氏名']).strip(),
                'has_companion': is_companion,
                'companion_name': companion_name,
            })
    return applicants


def run_lottery(applicants):
    """抽選を実行し、当選者リストと落選者リストを返す"""

    STATUS_PRIORITY = {'VIP': 0, '潜在': 1, '新規': 2}

    # 互いにペア申請しているペアをユニット化（スペース除去で名前を正規化して照合）
    def norm(s):
        return ''.join(s.split())

    norm_name_to_applicant = {norm(a['name']): a for a in applicants}
    paired_emails = set()
    pair_units = []  # [(primary, secondary), ...]

    for a in applicants:
        if a['email'] in paired_emails or not a['is_pair']:
            continue
        partner = norm_name_to_applicant.get(norm(a['pair_name']))
        if (partner and partner['is_pair']
                and norm(partner['pair_name']) == norm(a['name'])
                and partner['email'] not in paired_emails):
            pair_units.append((a, partner))
            paired_emails.add(a['email'])
            paired_emails.add(partner['email'])

    # ペア希望だが相手が見つからない人は落選扱い（ソロに入れない）
    unmatched_pair_emails = set()
    for a in applicants:
        if a['is_pair'] and a['email'] not in paired_emails:
            unmatched_pair_emails.add(a['email'])

    solos = [a for a in applicants if a['email'] not in paired_emails and a['email'] not in unmatched_pair_emails]

    # 枠ごとのステータス別残り人数 + 総残り人数
    slot_status_remaining = {sid: dict(ALLOCATION) for sid in SLOT_DEFS}
    slot_total = {sid: CAPACITY_PER_SLOT for sid in SLOT_DEFS}

    winners = []
    losers  = []
    assigned = set()
    checkin_id = 1

    def add_winner(applicant, sid):
        nonlocal checkin_id
        winners.append({
            'checkin_id': checkin_id,
            'name':       applicant['name'],
            'email':      applicant['email'],
            'status':     applicant['status'],
            'slot_id':    sid,
            'slot':       SLOT_DEFS[sid],
            'is_pair':    applicant['is_pair'],
            'pair_name':  applicant['pair_name'],
            'companion_name': applicant.get('companion_name', ''),
        })
        assigned.add(applicant['email'])
        checkin_id += 1

    # ソロをステータス別にシャッフル
    by_status = defaultdict(list)
    for a in solos:
        by_status[a['status']].append(a)
    for s in by_status:
        random.shuffle(by_status[s])

    # ペアユニットをステータス別（高い方優先）にシャッフル
    by_status_pairs = defaultdict(list)
    for (a, b) in pair_units:
        higher = min(a['status'], b['status'], key=lambda s: STATUS_PRIORITY[s])
        by_status_pairs[higher].append((a, b))
    for s in by_status_pairs:
        random.shuffle(by_status_pairs[s])

    # Pass1: VIP → 潜在 → 新規 の順で抽選
    for status in ['VIP', '潜在', '新規']:

        # ペアユニット（両者まとめて同じ枠に当選）
        for (a, b) in by_status_pairs[status]:
            if a['email'] in assigned or b['email'] in assigned:
                continue
            # 両者の共通希望枠で2人分の空きがある枠を探す
            common = a['preferred'] & b['preferred']
            if not common:
                common = a['preferred'] | b['preferred']
            available = [
                sid for sid in common
                if slot_status_remaining[sid][a['status']] >= 1
                and slot_status_remaining[sid][b['status']] >= 1
                and slot_total[sid] >= 2
            ]
            random.shuffle(available)
            if available:
                sid = available[0]
                slot_status_remaining[sid][a['status']] -= 1
                slot_status_remaining[sid][b['status']] -= 1
                slot_total[sid] -= 2
                add_winner(a, sid)
                add_winner(b, sid)

        # ソロ
        for applicant in by_status[status]:
            if applicant['email'] in assigned:
                continue
            available = [
                sid for sid in applicant['preferred']
                if slot_status_remaining[sid][status] >= 1
                and slot_total[sid] >= 1
            ]
            random.shuffle(available)
            if available:
                sid = available[0]
                slot_status_remaining[sid][status] -= 1
                slot_total[sid] -= 1
                add_winner(applicant, sid)

    # Pass2: ステータス枠の端数スペースを埋める
    for status in ['VIP', '潜在', '新規']:
        for applicant in by_status[status]:
            if applicant['email'] in assigned:
                continue
            available = [sid for sid in applicant['preferred'] if slot_total[sid] >= 1]
            random.shuffle(available)
            if available:
                slot_total[available[0]] -= 1
                add_winner(applicant, available[0])

    # 落選（ペアユニットの片方が漏れた場合も両者落選）
    for a in applicants:
        if a['email'] not in assigned:
            losers.append({
                'name':      a['name'],
                'email':     a['email'],
                'status':    a['status'],
                'preferred': a['slot_str'],
            })

    return winners, losers


def save_results(winners, losers, applicants):
    """Excelファイルに出力（全応募者・当選者・枠別シート）"""
    xlsx_path = OUTPUT_DIR + 'lottery_result.xlsx'
    wb = openpyxl.Workbook()

    # 色定義
    fill_header  = PatternFill('solid', fgColor='2C3E50')
    fill_vip     = PatternFill('solid', fgColor='F9E79F')
    fill_latent  = PatternFill('solid', fgColor='D5F5E3')
    fill_new     = PatternFill('solid', fgColor='FFFFFF')
    fill_winner  = PatternFill('solid', fgColor='AED6F1')
    font_header  = Font(color='FFFFFF', bold=True)

    def set_header(ws, cols):
        for i, col in enumerate(cols, 1):
            c = ws.cell(1, i, col)
            c.fill = fill_header
            c.font = font_header
            c.alignment = Alignment(horizontal='center')

    # --- Sheet1: 全応募者一覧 ---
    ws1 = wb.active
    ws1.title = '全応募者'
    set_header(ws1, ['No', '氏名', 'メールアドレス', 'ステータス', '希望日時', '結果', 'チェックインID', '参加枠'])

    winner_map = {w['email']: w for w in winners}
    status_fill = {'VIP': fill_vip, '潜在': fill_latent, '新規': fill_new}

    for i, a in enumerate(applicants, 1):
        w = winner_map.get(a['email'])
        result     = '当選' if w else '落選'
        checkin_id = w['checkin_id'] if w else ''
        slot       = w['slot'] if w else ''
        row = [i, a['name'], a['email'], a['status'], a['slot_str'], result, checkin_id, slot]
        ws1.append(row)
        fill = fill_winner if w else status_fill.get(a['status'], fill_new)
        for col in range(1, 9):
            ws1.cell(i + 1, col).fill = fill

    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['H'].width = 25

    # --- Sheet2: 当選者（メールマージ形式）---
    ws2 = wb.create_sheet('当選者一覧')
    headers2 = ['枠', 'Slot', 'ID', 'Name', '当選枠', '小分け', 'Email',
                'Column H', 'Column I', 'Column J', 'Column K', 'Column L',
                'Column M', 'Column N', 'Column O', 'Column P', 'Column Q', 'Column R',
                'Merge status']
    set_header(ws2, headers2)

    by_slot = defaultdict(list)
    for w in winners:
        by_slot[w['slot_id']].append(w)

    slot_num = 0
    for sid in sorted(SLOT_DEFS.keys()):
        slot_winners = sorted(by_slot[sid], key=lambda x: x['checkin_id'])
        if not slot_winners:
            continue
        slot_num += 1
        for w in slot_winners:
            fill = {'VIP': fill_vip, '潜在': fill_latent, '新規': fill_new}.get(w['status'], fill_new)
            row = [slot_num, SLOT_DEFS[sid], w['checkin_id'], w['name'], w['status'], '', w['email'],
                   '', '', '', '', '', '', '', '', '', '', '', '']
            ws2.append(row)
            for col in range(1, len(headers2) + 1):
                ws2.cell(ws2.max_row, col).fill = fill
        # 枠の間に空行
        ws2.append([''] * len(headers2))

    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['G'].width = 30

    # ペアカラーパレット（ステータス色とは別の鮮やかな色）
    PAIR_PALETTE = [
        'FFB3BA', 'BAE1FF', 'BAFFC9', 'FFE4B3', 'E8BAFF',
        'B3FFE8', 'FFB3E6', 'C9BAFF', 'FFFFBA', 'FFD1BA',
    ]

    def make_pair_color_map(slot_winners):
        """is_pair=True の全員に色を付ける。互いに名前を記載したペアは同色。"""
        def norm(s): return ''.join(s.split())
        norm_name_to_winner = {norm(w['name']): w for w in slot_winners}
        pair_group = {}   # email -> group_id
        group_count = 0
        for w in slot_winners:
            if not w['is_pair']:
                continue
            if w['email'] in pair_group:
                continue
            partner = norm_name_to_winner.get(norm(w['pair_name']))
            if partner and partner['email'] not in pair_group:
                pair_group[w['email']] = group_count
                pair_group[partner['email']] = group_count
            elif partner and partner['email'] in pair_group:
                pair_group[w['email']] = pair_group[partner['email']]
            else:
                pair_group[w['email']] = group_count
            group_count += 1
        return {
            email: PatternFill('solid', fgColor=PAIR_PALETTE[gid % len(PAIR_PALETTE)])
            for email, gid in pair_group.items()
        }

    fill_white = PatternFill('solid', fgColor='FFFFFF')

    def make_day_sheet(title, slot_ids):
        ws = wb.create_sheet(title)
        set_header(ws, ['参加枠', 'チェックインID', '氏名', 'ペア参加', 'メールアドレス', 'ステータス'])
        for sid in slot_ids:
            slot_winners = sorted(by_slot[sid], key=lambda x: x['checkin_id'])
            pair_color_map = make_pair_color_map(slot_winners)
            for w in slot_winners:
                pair_display = 'あり' if w['is_pair'] else 'なし'
                fill = pair_color_map.get(w['email'], fill_white)
                ws.append([SLOT_DEFS[sid], w['checkin_id'], w['name'], pair_display, w['email'], w['status']])
                for col in range(1, 7):
                    ws.cell(ws.max_row, col).fill = fill
            ws.append([''] * 6)  # 枠間の空行
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30

    # --- Sheet3: 5月9日 参加者 ---
    make_day_sheet('5月9日 参加者', range(1, 17))

    # --- Sheet4: 5月10日 参加者 ---
    make_day_sheet('5月10日 参加者', range(17, 33))

    # --- Sheet4: 落選者一覧 ---
    ws4 = wb.create_sheet('落選者一覧')
    set_header(ws4, ['氏名', 'メールアドレス', 'ステータス', '希望日時'])
    fill_loser = PatternFill('solid', fgColor='F2F3F4')
    for l in losers:
        ws4.append([l['name'], l['email'], l['status'], l['preferred']])
        for col in range(1, 5):
            ws4.cell(ws4.max_row, col).fill = fill_loser
    ws4.column_dimensions['B'].width = 30
    ws4.column_dimensions['D'].width = 20

    wb.save(xlsx_path)
    print(f'Excelファイル出力 → {xlsx_path}')
    print(f'  当選者: {len(winners)}名 / 落選者: {len(losers)}名')

    # Google Driveに上書きアップロード
    upload_to_drive(xlsx_path)


def upload_to_drive(xlsx_path):
    """Google Driveの固定ファイルに上書きアップロード"""
    import requests, subprocess, warnings
    warnings.filterwarnings('ignore')

    try:
        with open('/Users/ogawadaiki/.clasprc.json') as f:
            import json
            d = json.load(f)
        t = list(d['tokens'].values())[0]
        resp = requests.post('https://oauth2.googleapis.com/token', data={
            'client_id': t['client_id'],
            'client_secret': t['client_secret'],
            'refresh_token': t['refresh_token'],
            'grant_type': 'refresh_token'
        })
        token = resp.json()['access_token']

        with open(xlsx_path, 'rb') as f:
            file_data = f.read()

        resp = requests.patch(
            f'https://www.googleapis.com/upload/drive/v3/files/{GDRIVE_FILE_ID}?uploadType=media&fields=id,webViewLink',
            headers={
                'Authorization': f'Bearer {token}',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            },
            data=file_data
        )
        if resp.status_code == 200:
            print(f'Google Drive更新完了 → https://docs.google.com/spreadsheets/d/{GDRIVE_FILE_ID}/edit')
        else:
            print(f'Drive更新失敗: {resp.status_code} {resp.text[:200]}')
    except Exception as e:
        print(f'Drive更新エラー: {e}')


def print_summary(winners, losers, applicants):
    """集計サマリーを表示"""
    print('\n===== 抽選結果サマリー =====')
    print(f'総応募者数: {len(applicants)}名')
    print(f'当選: {len(winners)}名 / 落選: {len(losers)}名')

    # ステータス別集計
    print('\n--- ステータス別当選数 ---')
    for status in ['VIP', '潜在', '新規']:
        w = sum(1 for x in winners if x['status'] == status)
        l = sum(1 for x in losers  if x['status'] == status)
        print(f'  {status}: 当選{w}名 / 落選{l}名')

    # 枠別集計（実人数 = 当選者 + ペア）
    print('\n--- 枠別参加人数 ---')
    slot_winners_count = defaultdict(int)
    slot_people_count  = defaultdict(int)
    for w in winners:
        slot_winners_count[w['slot']] += 1
        slot_people_count[w['slot']]  += 2 if w['is_pair'] else 1
    for sid in sorted(SLOT_DEFS.keys()):
        name = SLOT_DEFS[sid]
        people = slot_people_count.get(name, 0)
        print(f'  {name}: {people}/{CAPACITY_PER_SLOT}名')


if __name__ == '__main__':
    print('Shopifyデータを読み込み中...')
    shopify = load_shopify(SHOPIFY_CSV)
    print(f'  顧客数: {len(shopify)}名')

    print('応募データを読み込み中...')
    applicants = load_applicants(PAPERFORM_CSV, shopify)
    print(f'  応募者数: {len(applicants)}名')

    # ステータス別集計
    for status in ['VIP', '潜在', '新規']:
        count = sum(1 for a in applicants if a['status'] == status)
        print(f'  {status}: {count}名')

    print('\n抽選実行中...')
    random.seed()  # 毎回ランダムなシードで実行
    winners, losers = run_lottery(applicants)

    # チェックインIDを枠順（5/9 1〜16 → 5/10 1〜16）・枠内順で振り直す
    by_slot = defaultdict(list)
    for w in winners:
        by_slot[w['slot_id']].append(w)
    new_id = 1
    for sid in sorted(SLOT_DEFS.keys()):
        for w in sorted(by_slot[sid], key=lambda x: x['checkin_id']):
            w['checkin_id'] = new_id
            new_id += 1

    print_summary(winners, losers, applicants)
    save_results(winners, losers, applicants)
    print('\n完了！')
